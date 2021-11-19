from openpyxl import load_workbook
from time import sleep
import sys

# Data Initialisation
path_in = ''
path_out = ''
wb = load_workbook(path_in)
ws = wb.active
bottomline = 12


# Transform to markdown
for row in range(2,bottomline):
    row = str(row)
    date = str(ws['C'+row].value).split(' 00:',1)[0] + ':\n'
    body = date + str(ws['D'+row].value) + '\n'
    for col in 'GHIJK':
        url = str(ws[col+row].value).split('=',1)[0]
        if len(url)>0:
            body = body + '![](' + url + '=s1000' + ')' + '\n'
            sleep(0.3)
        else:continue
    body = body + '\n---\n'
    print(body)
    with open(path_out,"a")as f:
        f.writelines(body)
        sleep(0.3)
