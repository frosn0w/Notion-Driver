import requests
from openpyxl import load_workbook

# Block ID
block_id = 'e4d978fcfaed4c308a3a6b7d7b9324dc'
# Notion Initialisation
token = ''
notion_version = '2021-08-16'
block_append_url = 'https://api.notion.com/v1/blocks/'+block_id+'/children'
headers={
    "Authorization":token,
    "Notion-Version":notion_version,
    "Content-Type":"application/json",
    }

# Data Initialisation
wb_path = '/Users/kuangbin/Desktop/bei-le-si.xlsx'
wb = load_workbook(wb_path)
ws = wb.active


# Post to Notion
print(block_append_url)
for i in range(2,30):
    num = str(i)
    img1 = str(ws['G'+num].value)+'\n'+'\n'
    img2 = str(ws['H'+num].value)+'\n'+'\n'
    img3 = str(ws['I'+num].value)+'\n'+'\n'
    img4 = str(ws['J'+num].value)+'\n'+'\n'
    img5 = str(ws['K'+num].value)+'\n'+'\n'
    body = '---\n' + str(ws['C'+num].value)+'：\n'+str(ws['D'+num].value)+img1+img2+img3+img4+img5
    data = {
        "children":
        [
            {
                "object": "block",
                "type": "paragraph",
                "paragraph":
                {
                    "text":
                    [
                        {
                            "type": "text",
                            "text": {"content": body}
                        }
                    ]
                }
            }
        ]
    }
    res = requests.patch(block_append_url, headers=headers, json=data)
    print(res.status_code)
    print(res.text)
